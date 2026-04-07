import os
import csv
import argparse
from datetime import datetime
from dotenv import load_dotenv
from elasticsearch import Elasticsearch
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Load environment
load_dotenv()

ES_URL = os.getenv("ES_URL")
ES_API_KEY = os.getenv("ES_API_KEY")
ES_INDEX = os.getenv("ES_INDEX", "*")
VERIFY_CERTS = os.getenv("VERIFY_CERTS", "false").lower() == "true"

# Basic validation
missing = [name for name, val in (("ES_URL", ES_URL), ("ES_API_KEY", ES_API_KEY)) if not val]
if missing:
    print(f"[!] Missing required environment variables: {', '.join(missing)}\nSee .env.example for expected values.")
    raise SystemExit(1)

# Fields to export
FIELDS = [
    "@timestamp",
    "event.module",
    "event.dataset",
    "event.type",
    "event.action",
    "message",
    "host.id",
    "host.name",
    "host.hostname",
    "host.os.name",
    "host.os.type",
    "agent.name",
    "agent.type",
    "user.name",
    "user.id",
    "user.domain",
    "process.name",
    "process.executable",
    "process.command_line",
    "process.args",
    "process.pid",
    "process.working_directory",
    "process.parent.name",
    "process.parent.pid",
    "process.parent.command_line",
    "process.parent.executable",
    "process.parent.hash.sha256",
    "process.hash.sha256",
    "process.hash.md5",
    "process.hash.sha1",
    "process.code_signature.exists",
    "process.code_signature.valid",
    "process.code_signature.trusted",
    "process.code_signature.subject_name",
    "process.code_signature.status",
    "process.pe.original_file_name",
    "process.entity_id",
    "process.start",
    "file.path",
    "file.hash.sha256",
    "file.code_signature.exists",
    "file.code_signature.trusted",
    "file.extension",
    "dll.code_signature.exists",
    "dll.code_signature.trusted",
    "dll.path",
    "event.code",
    "source.ip",
    "destination.ip",
    "destination.port",
    "network.transport",
    "network.direction",
    "process.parent.code_signature.exists",
    "process.parent.code_signature.status",
    "process.parent.code_signature.subject_name",
    "process.parent.code_signature.trusted",
]

SOURCE_FIELDS = FIELDS.copy()

# Fields for per-rule classification sheet (leading column shows which rule(s) matched)
RULES_FIELDS = ["matched_rules"] + SOURCE_FIELDS.copy()

# Fields for Rule 1 sheet (includes validation columns)
RULE1_FIELDS = [
    "Val1 User Context",
    "Val2 Parent/Child Process",
] + SOURCE_FIELDS.copy()

# Fields for Rule 2 sheet (includes validation columns)
RULE2_FIELDS = [
    "Val1 Parent/Child Process",
    "Val2 DLL Signature",
] + SOURCE_FIELDS.copy()

# Fields for Rule 3 sheet (includes validation columns)
RULE3_FIELDS = [
    "Val1 DLL Load Presence",
] + SOURCE_FIELDS.copy()

# Fields for Rule 4 sheet (includes validation columns)
RULE4_FIELDS = [
    "Val1 File Signature Check",
    "Val2 DLL Load Detection",
] + SOURCE_FIELDS.copy()

# Fields for Rule 5 sheet (includes validation columns)
RULE5_FIELDS = [
    "Val1 Network Type",
    "Val2 Network Timing",
] + SOURCE_FIELDS.copy()

# Fields for Rule 6 sheet (includes validation columns)
RULE6_FIELDS = [
    "Val1 User Context",
    "Val2 Network Connection",
    "Val3 Parent Signature",
] + SOURCE_FIELDS.copy()

# Fields for Rule 7 sheet (includes validation columns)
RULE7_FIELDS = [
    "Val1 DLL File Location",
    "Val2 Network Activity",
] + SOURCE_FIELDS.copy()

# Fields for Rule 8 sheet (includes validation columns)
RULE8_FIELDS = [
    "Val1 Child Process Risk",
    "Val2 DLL Path in Command",
] + SOURCE_FIELDS.copy()

# Fields for Rule 9 sheet (includes validation columns)
RULE9_FIELDS = [
    "Val1 Signature Publisher",
    "Val2 File Path",
] + SOURCE_FIELDS.copy()

# Fields for Rule 10 sheet (includes validation columns)
RULE10_FIELDS = [
    "Val1 Parent Process",
    "Val2 Network Context",
] + SOURCE_FIELDS.copy()

# Short descriptions for each rule to include in classification output
RULE_DESCRIPTIONS = {
    1: "Scriptlet/Remote execution (/i: + scrobj.dll/URL/UNC/.sct)",
    2: "Non-system path DLL",
    3: "Non-standard extension",
    4: "Double-extension masquerade",
    5: "Network Activity/Connections",
    6: "Parent not installer/suspicious parent",
    7: "DLL unsigned",
    8: "Child processes spawned by regsvr32",
    9: "Process (regsvr32) unsigned or untrusted hash",
    10: "SMB Share Remote Execution (UNC path + SMB network)",
    11: "Renamed regsvr32",
    12: "DLL from user-writable path",
}

# PROCESS TREE - Suspicious parent processes
SUSPICIOUS_PARENTS = [
    "powershell.exe",
    "pwsh.exe",
    "cmd.exe",
    "wscript.exe",
    "cscript.exe",
    "mshta.exe",
    "rundll32.exe",
    "wmic.exe",
    "wmiprvse.exe",  # WMI Provider Host (remote execution)
    "winrs.exe",     # WinRM remote shell
    "wsmprovhost.exe",  # WinRM provider host
    "psexec.exe",    # Sysinternals PsExec
    "psexesvc.exe",  # PsExec service
    "paexec.exe",    # PAExec (PsExec-like)
    "schtasks.exe",  # Scheduled tasks (remote execution)
    "at.exe",        # AT scheduler (legacy)
    "sc.exe",        # Service control (remote service exec)
    "reg.exe",       # Remote registry ops
    "net.exe",       # Net commands
    "net1.exe",      # Net fallback
    "winrm.cmd",    # WinRM client wrapper
    "wmiadap.exe",  # WMI adapter
    "mofcomp.exe",  # WMI MOF compiler
    "wbemtest.exe", # WMI tester
    "powershell_ise.exe",  # PowerShell ISE
    "msbuild.exe",  # LOLBin build execution
    "installutil.exe", # LOLBin installer
    "regasm.exe",   # .NET registration
    "regsvcs.exe",  # .NET services
    "cmdkey.exe",   # Credential manager
    "rdesktop.exe", # RDP client (if present)
    "mstsc.exe",    # RDP client
    "mimikatz.exe", # Post-exploitation tool (if present)
    "procdump.exe", # Sysinternals dump tool
    "wmiexec.exe",  # Impacket WMIExec (if present)
    "smbexec.exe",  # Impacket SMBExec (if present)
    "atexec.exe",   # Impacket ATExec (if present)
    "dcomexec.exe", # Impacket DCOMExec (if present)
    "psremoting.exe", # PowerShell remoting helper (if present)
    "psexec64.exe", # PsExec 64-bit
    "sexec.exe",    # PsExec variants
    "remote.exe",   # Generic remote tools (if present)
    "agent.exe",    # Remote agent (if present)
    "teamviewer.exe",
    "anydesk.exe",
    "screenconnect.exe",
    "splashtop.exe",
    "logmein.exe",
    "excel.exe",
    "winword.exe",
    "outlook.exe",
    "powerpnt.exe",
    "chrome.exe",
    "firefox.exe",
    "iexplore.exe",
    "msedge.exe",
    "explorer.exe",  # Suspicious when spawning regsvr32 directly
]

# LEGITIMATE PARENTS - Known safe parent processes
LEGITIMATE_PARENTS = [
    # Windows installer/service processes
    "services.exe",
    "svchost.exe",
    "msiexec.exe",
    "setup.exe",
    "installer.exe",
    "installutil.exe",
    "dllhost.exe",
    "mmc.exe",
    "taskmgr.exe",
    
    # Windows system/update processes
    "trustedinstaller.exe",      # Windows Update & Feature installation
    "spoolsv.exe",               # Print spooler (registers printer drivers)
    "audiodg.exe",               # Audio Device Graph (audio/media drivers)
    "searchindexer.exe",         # Windows Search service
    "winsatmediarating.exe",     # Windows Media Rating service
    "conhost.exe",               # Console host
    "lsass.exe",                 # Local Security Authority
    
    # Common software installers
    "apphelpcapv1.exe",          # Application help handler
    "apphelpercap.exe",          # HP application helper
    "hpsaapplauncher.exe",       # HP application launcher
    
    # Runtime/framework processes
    "dotnet.exe",                # .NET runtime
    "java.exe",                  # Java runtime
    "javaw.exe",                 # Java runtime (windowed)
]

# Time ranges
TIME_RANGES = {
    "1": ("10 seconds", "now-10s"),
    "2": ("1 hour", "now-1h"),
    "3": ("6 hours", "now-6h"),
    "4": ("12 hours", "now-12h"),
    "5": ("24 hours (1 day)", "now-24h"),
    "6": ("48 hours (2 days)", "now-48h"),
    "7": ("72 hours (3 days)", "now-72h"),
    "8": ("1 week", "now-168h"),
    "9": ("2 weeks", "now-336h"),
    "10": ("1 month (30 days)", "now-720h"),
    "11": ("3 months", "now-2160h"),
    "12": ("6 months", "now-4320h"),
    "13": ("1 year", "now-8760h"),
    "14": ("All time", None),
}

def get_value(src, path):
    """Traverse nested dicts using dotted path."""
    parts = path.split(".") if path else []
    cur = src
    try:
        for p in parts:
            if cur is None:
                return ""
            if isinstance(cur, list):
                values = []
                for item in cur:
                    if isinstance(item, dict) and p in item:
                        values.append(item[p])
                    else:
                        values.append(item)
                cur = values
                continue
            if isinstance(cur, dict):
                if p in cur:
                    cur = cur[p]
                    continue
                alt = p.replace("-", "_")
                if alt in cur:
                    cur = cur[alt]
                    continue
                return ""
            return ""
        if isinstance(cur, list):
            return "; ".join([str(x) for x in cur])
        if cur is None:
            return ""
        return str(cur)
    except Exception:
        return ""

def serialize_cell(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        try:
            import json
            return json.dumps(v, ensure_ascii=False)
        except Exception:
            return str(v)
    return str(v)

def check_user_privilege(user_name, user_id=None):
    """
    Determine user privilege level based on user ID (SID) only.
    Returns: 'system', 'admin', or 'unknown'
    Note: user.name is NOT checked (can be easily spoofed)
    """
    # Check user ID (SID) only - cannot be easily spoofed
    if user_id:
        user_id_lower = str(user_id).lower()
        
        # Well-known SYSTEM SIDs
        if any(sys_sid in user_id_lower for sys_sid in [
            "s-1-5-18",  # LocalSystem
            "s-1-5-19",  # LocalService
            "s-1-5-20",  # NetworkService
        ]):
            return "system"
        
        # Administrator SIDs (RID 500 or 544 for Administrators group)
        if "s-1-5-21" in user_id_lower and ("-500" in user_id_lower or "-544" in user_id_lower):
            return "admin"
        
        # Domain Admins (RID 512), Enterprise Admins (RID 519)
        if "s-1-5-21" in user_id_lower and ("-512" in user_id_lower or "-519" in user_id_lower):
            return "admin"
    
    # No SID available - return unknown (do not fallback to user.name which can be spoofed)
    return "unknown"

def validate_rule1_user_context(src):
    """
    VALIDATION 1 for Rule 1: Check user context by SID (user.id)
    - Normal user (no admin SID) → MALICIOUS
    - Admin/System user → SUSPICIOUS (need admin validation)
    
    Returns: "MALICIOUS" | "SUSPICIOUS - Reach out to admin" | "N/A"
    """
    user_id = get_value(src, "user.id")
    user_name = get_value(src, "user.name")
    
    privilege = check_user_privilege(user_name, user_id)
    
    if privilege == "unknown":
        return "MALICIOUS (Normal user - no admin SID found)"
    elif privilege in ["admin", "system"]:
        return f"SUSPICIOUS ({privilege} context - Reach out to admin for validation)"
    else:
        return "N/A"

def validate_rule1_parent_process(src):
    """
    VALIDATION 2 for Rule 1: Check for legitimate installer parent
    - No legit parent → MALICIOUS
    - Has legit parent → SUSPICIOUS (uncommon approach for modern installation)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    parent_name = get_value(src, "process.parent.name").lower()
    
    if not parent_name:
        return "MALICIOUS (No parent process information)"
    
    # Check if parent is a legitimate installer
    has_legit_parent = any(legit in parent_name for legit in LEGITIMATE_PARENTS)
    
    if has_legit_parent:
        return f"SUSPICIOUS (Legit parent: {parent_name} - Uncommon approach for modern installation)"
    else:
        return f"MALICIOUS (No legitimate installer parent - Current parent: {parent_name})"

def validate_rule2_parent_process(src):
    """
    VALIDATION 1 for Rule 2: Check parent process and process tree
    - If parent is NOT a legitimate installer → MALICIOUS
    - If process.parent.name contains "regsvr32" AND process.name exists → MALICIOUS (spawned child)
    - Otherwise → SUSPICIOUS
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    parent_name = get_value(src, "process.parent.name").lower()
    
    if not parent_name:
        return "MALICIOUS (No parent process information)"
    
    # Check if parent is a legitimate installer
    has_legit_parent = any(legit in parent_name for legit in LEGITIMATE_PARENTS)
    
    if not has_legit_parent:
        return f"MALICIOUS (Parent not legitimate installer - Current parent: {parent_name})"
    
    # Check if this event is a child process spawned by regsvr32
    # Correlate: if process.parent.name = regsvr32 AND process.name exists
    if "regsvr32" in parent_name:
        current_process_name = get_value(src, "process.name").lower()
        if current_process_name:
            return f"MALICIOUS (Spawned child process: {current_process_name} - Post-exploitation indicator)"
    
    return "SUSPICIOUS (Legit parent but unusual DLL loading from non-system path)"

def validate_rule2_dll_signature(src):
    """
    VALIDATION 2 for Rule 2: Check DLL signature status
    - If DLL is NOT signed (dll.code_signature.exists = false/no signature) → MALICIOUS
    - If DLL is signed → SUSPICIOUS (signed DLLs from user paths are still unusual)
    - If signature info not available → MALICIOUS (assume unsigned for safety)
    
    Returns: "MALICIOUS (Unsigned)" | "SUSPICIOUS (Signed)" | "N/A"
    """
    # Check if DLL has signature information from dll.code_signature
    dll_sig_exists = get_value(src, "dll.code_signature.exists")
    
    # Convert string values to boolean if necessary
    if isinstance(dll_sig_exists, str):
        dll_sig_exists = dll_sig_exists.lower() in ("true", "yes", "1")
    
    # If no signature exists → MALICIOUS
    if not dll_sig_exists:
        return "MALICIOUS (Unsigned DLL - No code signature)"
    
    # If signature exists → SUSPICIOUS (signed DLLs from user paths are still unusual)
    if dll_sig_exists:
        return "SUSPICIOUS (DLL is signed but loaded from non-system path)"
    
    # If signature information not available → MALICIOUS (assume unsigned for safety)
    return "MALICIOUS (No signature information available - Assume unsigned)"

def validate_rule3_no_dll_load(src):
    """
    VALIDATION 1 for Rule 3: Non-DLL extension AND no DLL load observed
    - If command line references non-DLL extension AND no DLL load indicators → MALICIOUS
    - If DLL load indicators are present → SUSPICIOUS (possible mixed activity)
    - If command line is missing → N/A

    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    cmdline = get_value(src, "process.command_line")
    if not cmdline:
        return "N/A (No command line)"

    cmdline_lower = cmdline.lower()
    import re
    # Extract extensions and treat anything other than .dll/.ax/.ocx as non-DLL
    exts = re.findall(r"\.([a-z0-9]{1,5})\b", cmdline_lower)
    has_non_dll_ext = any(ext not in ("dll", "ax", "ocx") for ext in exts)

    if not has_non_dll_ext:
        return "N/A (No non-DLL extension other than .dll/.ax/.ocx found)"

    event_category = get_value(src, "event.category")
    dll_path = get_value(src, "dll.path")
    dll_sig_exists = get_value(src, "dll.code_signature.exists")

    # Normalize event.category to string list or string
    event_category_lower = str(event_category).lower()

    # Indicators of DLL load
    has_dll_load = (
        "library" in event_category_lower or
        bool(dll_path) or
        (isinstance(dll_sig_exists, str) and dll_sig_exists.lower() in ("true", "yes", "1")) or
        (isinstance(dll_sig_exists, bool) and dll_sig_exists)
    )

    if not has_dll_load:
        return "MALICIOUS (Non-DLL extension + no DLL load observed)"

    return "SUSPICIOUS (Non-DLL extension but DLL load indicators present)"

def validate_rule4_signature_check(src):
    """
    VALIDATION 1 for Rule 4: Check signature based on event type
    - If event.category = library → check dll.code_signature
    - If not library event → check file.code_signature
    - Unsigned or untrusted → MALICIOUS
    - Signed and trusted → BENIGN

    Returns: "MALICIOUS" | "BENIGN" | "N/A"
    """
    event_category = get_value(src, "event.category")
    
    # Normalize event.category to string
    event_category_lower = str(event_category).lower()
    
    # Check if this is a library event
    is_library = "library" in event_category_lower
    
    if is_library:
        # Check DLL signature for library events
        sig_exists = get_value(src, "dll.code_signature.exists")
        sig_trusted = get_value(src, "dll.code_signature.trusted")
        sig_type = "DLL"
    else:
        # Check file signature for non-library events
        sig_exists = get_value(src, "file.code_signature.exists")
        sig_trusted = get_value(src, "file.code_signature.trusted")
        sig_type = "File"
    
    # Convert string values to boolean if necessary
    if isinstance(sig_exists, str):
        sig_exists = sig_exists.lower() in ("true", "yes", "1")
    if isinstance(sig_trusted, str):
        sig_trusted = sig_trusted.lower() in ("true", "yes", "1")
    
    # No signature → MALICIOUS
    if not sig_exists:
        return f"MALICIOUS (Unsigned {sig_type})"
    
    # Signature exists but untrusted → MALICIOUS
    if sig_exists and not sig_trusted:
        return f"MALICIOUS (Untrusted {sig_type} signature)"
    
    # Signed and trusted → BENIGN
    if sig_exists and sig_trusted:
        return f"BENIGN (Signed and trusted {sig_type})"
    
    return "N/A (No signature info)"

def validate_rule4_dll_load_detection(src):
    """
    VALIDATION 2 for Rule 4: Check if DLL load event exists with DLL name
    - If event.category = library AND dll.path exists → DLL LOADED (shows DLL path)
    - If event.category = library but no dll.path → SUSPICIOUS (library event without DLL info)
    - If not a library event → N/A

    Returns: "DLL LOADED: <path>" | "SUSPICIOUS" | "N/A"
    """
    event_category = get_value(src, "event.category")
    dll_path = get_value(src, "dll.path")
    dll_name = get_value(src, "dll.name")
    
    # Normalize event.category to string
    event_category_lower = str(event_category).lower()
    
    # Check if this is a library event
    if "library" not in event_category_lower:
        return "N/A (Not a library event)"
    
    # Library event detected - check for DLL information
    if dll_path:
        return f"DLL LOADED: {dll_path}"
    elif dll_name:
        return f"DLL LOADED: {dll_name}"
    else:
        return "SUSPICIOUS (Library event without DLL info)"

def validate_rule5_network_type(src):
    """
    VALIDATION 1 for Rule 5: Check if network connection is internal or external
    - If external network → MALICIOUS
    - If internal network → Check user privilege:
        - Normal user → MALICIOUS (suspicious internal communication)
        - Admin user → SUSPICIOUS (uncommon installation approach)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    dest_ip = get_value(src, "destination.ip")
    user_id = get_value(src, "user.id")
    user_name = get_value(src, "user.name")
    
    if not dest_ip:
        return "N/A (No destination IP)"
    
    # Check if destination IP is internal (RFC1918 private ranges)
    is_internal = (
        dest_ip.startswith("10.") or
        dest_ip.startswith("172.16.") or dest_ip.startswith("172.17.") or 
        dest_ip.startswith("172.18.") or dest_ip.startswith("172.19.") or
        dest_ip.startswith("172.20.") or dest_ip.startswith("172.21.") or
        dest_ip.startswith("172.22.") or dest_ip.startswith("172.23.") or
        dest_ip.startswith("172.24.") or dest_ip.startswith("172.25.") or
        dest_ip.startswith("172.26.") or dest_ip.startswith("172.27.") or
        dest_ip.startswith("172.28.") or dest_ip.startswith("172.29.") or
        dest_ip.startswith("172.30.") or dest_ip.startswith("172.31.") or
        dest_ip.startswith("192.168.") or
        dest_ip.startswith("127.") or  # Loopback
        dest_ip == "localhost"
    )
    
    if not is_internal:
        return "MALICIOUS (External network connection - C2 communication)"
    
    # Internal network - check user privilege
    privilege = check_user_privilege(user_name, user_id)
    
    if privilege == "admin" or privilege == "system":
        return "SUSPICIOUS (Admin user, internal network - Uncommon DLL Registration)"
    else:
        return "MALICIOUS (Normal user making internal network calls via regsvr32)"
    
def validate_rule5_network_timing(src):
    """
    VALIDATION 2 for Rule 5: Check timing between regsvr32 execution and network activity
    - If network activity occurs within X seconds (e.g., 5 seconds) of regsvr32 start → MALICIOUS
    - If timing is unclear or delayed → SUSPICIOUS
    - If no timing data → N/A
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    timestamp = get_value(src, "@timestamp")
    network_timestamp = get_value(src, "event.start") or timestamp  # Network event time
    process_start = get_value(src, "process.start")  # Preferred for correlation
    process_entity_id = get_value(src, "process.entity_id")
    
    if not (timestamp or process_start) or not network_timestamp:
        return "N/A (No timing data available)"
    
    base_time = process_start or timestamp
    base_label = "process.start" if process_start else "@timestamp"
    correlation = "strong" if process_entity_id else "weak"
    
    # Convert to datetime for comparison
    try:
        from datetime import datetime
        # Parse ISO format timestamps
        ts1 = datetime.fromisoformat(base_time.replace('Z', '+00:00'))
        ts2 = datetime.fromisoformat(network_timestamp.replace('Z', '+00:00'))
        
        time_diff = abs((ts2 - ts1).total_seconds())
        
        # If network activity within 5 seconds of regsvr32 execution → MALICIOUS
        if time_diff <= 5:
            return f"MALICIOUS (Network activity within {int(time_diff)}s of execution - Direct C2 | {correlation} correlation via {base_label})"
        # If within 30 seconds → SUSPICIOUS
        elif time_diff <= 30:
            return f"SUSPICIOUS (Network activity within {int(time_diff)}s - Possible C2 | {correlation} correlation via {base_label})"
        else:
            return f"SUSPICIOUS (Network activity {int(time_diff)}s after execution | {correlation} correlation via {base_label})"
    except Exception:
        return "SUSPICIOUS (Unable to parse timing - Assume suspicious)"

def validate_rule6_user_privilege(src):
    """
    VALIDATION 1 for Rule 6: Determine user privilege level
    - Admin/System user → SUSPICIOUS (uncommon but possible for deployment)
    - Normal user → MALICIOUS (users shouldn't register DLLs)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    user_id = get_value(src, "user.id")
    user_name = get_value(src, "user.name")
    
    if not user_id and not user_name:
        return "N/A (No user context)"
    
    privilege = check_user_privilege(user_name, user_id)
    
    if privilege == "admin":
        return f"SUSPICIOUS (Admin user - {user_name})"
    elif privilege == "system":
        return f"SUSPICIOUS (System user - {user_name})"
    else:
        return f"MALICIOUS (Normal user - {user_name})"

def validate_rule6_network_connection(src):
    """
    VALIDATION 2 for Rule 6: Check for network connections after regsvr32 execution
    - External network connection → MALICIOUS (C2 communication)
    - Internal network connection → SUSPICIOUS (lateral movement)
    - No network activity → BENIGN
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "BENIGN"
    """
    dest_ip = get_value(src, "destination.ip")
    dest_port = get_value(src, "destination.port")
    net_direction = get_value(src, "network.direction")
    
    if not dest_ip:
        return "No network activity detected"
    
    is_internal = (
        dest_ip.startswith("10.") or
        dest_ip.startswith("172.16.") or dest_ip.startswith("172.17.") or 
        dest_ip.startswith("172.18.") or dest_ip.startswith("172.19.") or
        dest_ip.startswith("172.20.") or dest_ip.startswith("172.21.") or
        dest_ip.startswith("172.22.") or dest_ip.startswith("172.23.") or
        dest_ip.startswith("172.24.") or dest_ip.startswith("172.25.") or
        dest_ip.startswith("172.26.") or dest_ip.startswith("172.27.") or
        dest_ip.startswith("172.28.") or dest_ip.startswith("172.29.") or
        dest_ip.startswith("172.30.") or dest_ip.startswith("172.31.") or
        dest_ip.startswith("192.168.") or
        dest_ip.startswith("127.")
    )
    
    port_info = f":{dest_port}" if dest_port else ""
    direction_info = f" ({net_direction})" if net_direction else ""
    
    if not is_internal:
        return f"MALICIOUS (External connection - C2: {dest_ip}{port_info}{direction_info})"
    else:
        return f"SUSPICIOUS (Internal connection - Lateral movement: {dest_ip}{port_info}{direction_info})"

def validate_rule6_parent_signature(src):
    """
    VALIDATION 3 for Rule 6: Check parent process signature (detect masquerading/tampering)
    - Parent NOT Microsoft-signed → MALICIOUS (masquerading or tampered system binary)
    - Parent Microsoft-signed but untrusted → SUSPICIOUS (could be legitimate but rare)
    - Parent Microsoft-signed and trusted → BENIGN (legitimate system process)
    - No signature info → SUSPICIOUS (cannot verify parent legitimacy)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "BENIGN" | "N/A"
    """
    parent_sig_exists = get_value(src, "process.parent.code_signature.exists")
    parent_sig_trusted = get_value(src, "process.parent.code_signature.trusted")
    parent_sig_subject = get_value(src, "process.parent.code_signature.subject_name")
    parent_executable = get_value(src, "process.parent.executable")
    parent_name = get_value(src, "process.parent.name")
    
    # Convert string values to boolean if necessary
    if isinstance(parent_sig_exists, str):
        parent_sig_exists = parent_sig_exists.lower() in ("true", "yes", "1")
    if isinstance(parent_sig_trusted, str):
        parent_sig_trusted = parent_sig_trusted.lower() in ("true", "yes", "1")
    
    if not parent_sig_exists:
        # Parent has no signature - could be masquerading
        return f"MALICIOUS (Unsigned parent - {parent_name} from {parent_executable})"
    
    # Parent is signed but check if trusted
    if parent_sig_exists and not parent_sig_trusted:
        # Signed but untrusted - suspicious
        return f"SUSPICIOUS (Untrusted/Non-Microsoft signature - {parent_sig_subject})"
    
    # Parent is signed and trusted - check if it's Microsoft
    if parent_sig_exists and parent_sig_trusted:
        if parent_sig_subject and "microsoft" in parent_sig_subject.lower():
            return f"Microsoft-signed trusted parent - {parent_name})"
        else:
            return f"SUSPICIOUS (Trusted but non-Microsoft parent - {parent_sig_subject})"
    
    # No signature information available
    return "SUSPICIOUS (Cannot verify parent process signature)"

def validate_rule7_dll_file_location(src):
    """
    VALIDATION 1 for Rule 7: Check DLL file location (dll.path)
    - Legitimate system paths (system32, syswow64, program files, program files x86) → BENIGN
    - Anything else → MALICIOUS
    
    Returns: "MALICIOUS" | "BENIGN" | "N/A"
    """
    dll_path = get_value(src, "dll.path")
    
    if not dll_path:
        return "N/A (No DLL path information)"
    
    dll_path_lower = dll_path.lower()
    
    # Legitimate system paths (case-insensitive check)
    legit_paths = [
        "*\\\\windows\\\\system32\\\\*",
        "*\\\\windows\\\\syswow64\\\\*",
        "*\\\\program files\\\\*",
        "*\\\\program files (x86)\\\\*",
        "*\\\\Windows\\\\System32\\\\*",  # Case variations
        "*\\\\Windows\\\\SysWOW64\\\\*",
        "*\\\\Program Files\\\\*",
        "*\\\\Program Files (x86)\\\\*"
    ]
    
    for path in legit_paths:
        # Remove wildcards for substring matching in validation
        path_pattern = path.strip('*').replace('\\\\', '\\\\')
        if path_pattern.lower() in dll_path_lower:
            return f"BENIGN (DLL from legitimate system path)"
    
    # Everything else is malicious
    return f"MALICIOUS (DLL from non-system location)"

def validate_rule7_network_activity_presence(src):
    """
    VALIDATION 2 for Rule 7: Check if network activity occurs after DLL registration
    - Network activity within 30 seconds after execution (external or internal) → MALICIOUS
    - Detects both C2 callbacks (external) and lateral movement (internal)
    - Network activity detected but timing unclear → SUSPICIOUS
    - No network activity → BENIGN
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "BENIGN" | "N/A"
    """
    process_start = get_value(src, "process.start")
    timestamp = get_value(src, "@timestamp")
    event_start = get_value(src, "event.start") or timestamp
    dest_ip = get_value(src, "destination.ip")
    process_entity_id = get_value(src, "process.entity_id")
    
    # No network activity
    if not dest_ip:
        return "BENIGN (No network activity after DLL registration)"
    
    # Network activity present but no timing data
    if not (process_start or timestamp) or not event_start:
        correlation = "strong" if process_entity_id else "weak"
        return f"SUSPICIOUS (Network activity detected but timing unclear | {correlation} correlation via entity_id: {dest_ip})"
    
    # Calculate timing between process start and network event
    base_time = process_start or timestamp
    try:
        from datetime import datetime
        ts1 = datetime.fromisoformat(base_time.replace('Z', '+00:00'))
        ts2 = datetime.fromisoformat(event_start.replace('Z', '+00:00'))
        
        time_diff = abs((ts2 - ts1).total_seconds())
        correlation = "strong" if process_entity_id else "weak"
        
        # Classify IP type
        is_internal = dest_ip.startswith(("10.", "172.16.", "172.17.", "172.18.", "172.19.", "172.20.", "172.21.", "172.22.", "172.23.", "172.24.", "172.25.", "172.26.", "172.27.", "172.28.", "172.29.", "172.30.", "172.31.", "192.168.", "127."))
        ip_type = "Internal (Lateral Movement)" if is_internal else "External (C2 Callback)"
        
        # Network activity within 30 seconds → MALICIOUS
        if time_diff <= 30:
            return f"MALICIOUS (Network within {int(time_diff)}s - {ip_type} | {correlation} correlation: {dest_ip})"
        # Network activity 30-300 seconds → SUSPICIOUS
        elif time_diff <= 300:
            return f"SUSPICIOUS (Network {int(time_diff)}s after execution - {ip_type} | {correlation} correlation: {dest_ip})"
        else:
            return f"SUSPICIOUS (Network {int(time_diff)}s after execution - delayed {ip_type} | {correlation} correlation: {dest_ip})"
    except Exception:
        correlation = "strong" if process_entity_id else "weak"
        return f"SUSPICIOUS (Unable to parse timing - network to {dest_ip} | {correlation} correlation)"
   
def validate_rule8_child_process(src):
    """
    VALIDATION 2 for Rule 8: Check child process type and command line
    - High-risk process (shells/scripts) + suspicious commands → MALICIOUS
    - System tools + enumeration/persistence → MALICIOUS
    - Regsvr32 spawning regsvr32 → MALICIOUS (self-replication)
    - Other processes → SUSPICIOUS
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    child_name = get_value(src, "process.name")
    child_cmd = get_value(src, "process.command_line")
    
    if not child_name:
        return "N/A (No child process info)"
    
    child_name = child_name.lower()
    child_cmd = child_cmd.lower() if child_cmd else ""
    
    # High-risk shells and scripting hosts
    shells = ["cmd.exe", "powershell.exe", "pwsh.exe", "wscript.exe", "cscript.exe", "mshta.exe"]
    # System/admin tools often used for malicious activity
    admin_tools = ["net.exe", "net1.exe", "sc.exe", "reg.exe", "schtasks.exe", "certutil.exe", 
                   "bitsadmin.exe", "wmic.exe", "whoami.exe", "nltest.exe", "tasklist.exe"]
    # LOLBin chaining
    lolbins = ["rundll32.exe", "regsvr32.exe", "msiexec.exe", "installutil.exe"]
    
    # Check for self-spawning (regsvr32 → regsvr32)
    if "regsvr32" in child_name:
        return f"SUSPICIOUS (regsvr32 spawning {child_name})"
    
    # Check for shell/script processes
    if any(shell in child_name for shell in shells):
        # Check command line for suspicious indicators
        if any(indicator in child_cmd for indicator in ["http://", "https://", "downloadstring", "invoke-", 
                                                          "iex", "base64", "-enc", "encodedcommand",
                                                          "bypass", "hidden", "noprofile"]):
            return f"MALICIOUS (Shell + Suspicious command - Likely C2/download) - {child_name}"
        return f"MALICIOUS (Shell/Script execution from regsvr32) - {child_name}"
    
    # Check for admin/system tools
    if any(tool in child_name for tool in admin_tools):
        if any(indicator in child_cmd for indicator in ["user", "service", "schedule", "registry", 
                                                          "download", "transfer", "process"]):
            return f"MALICIOUS (System tool + Enumeration/Persistence) - {child_name}"
        return f"MALICIOUS (Admin tool execution from regsvr32) - {child_name}"
    
    # Check for LOLBin chaining
    if any(lolbin in child_name for lolbin in lolbins):
        return f"MALICIOUS (LOLBin chaining) - {child_name}"
    
    # Any other child process is still suspicious
    return f"SUSPICIOUS (Unusual child process: {child_name})"

def validate_rule8_dll_path_in_command(src):
    """
    VALIDATION 2 for Rule 8: Check DLL path in both process and parent command lines
    - Checks both process.command_line and process.parent.command_line
    - Legitimate system paths (system32, syswow64, program files, program files x86) → BENIGN
    - Anything else → MALICIOUS
    
    Returns: "MALICIOUS" | "BENIGN" | "N/A"
    """
    process_cmd = get_value(src, "process.command_line")
    parent_cmd = get_value(src, "process.parent.command_line")

    if not process_cmd and not parent_cmd:
        return "N/A (No command line)"

    # Legitimate system paths (case-insensitive wildcard patterns)
    legit_paths = [
        "*\\windows\\system32\\*",
        "*\\windows\\syswow64\\*",
        "*\\program files\\*",
        "*\\program files (x86)\\*",
    ]

    combined_cmd = f"{process_cmd or ''} {parent_cmd or ''}"

    import fnmatch
    combined_cmd_lower = combined_cmd.lower()
    if any(fnmatch.fnmatch(combined_cmd_lower, pattern.lower()) for pattern in legit_paths):
        return "BENIGN (Legitimate path found in command line)"

    return "MALICIOUS (No legitimate path found in command line)"

def validate_rule9_signature_publisher(src):
    """
    VALIDATION 1 for Rule 9: Check code signature publisher
    - Publisher is NOT Microsoft → MALICIOUS (Unsigned or third-party signed)
    - Publisher is unknown/empty → SUSPICIOUS (No publisher info)
    - Publisher is Microsoft Corporation → BENIGN (Microsoft-signed legitimate binary)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "BENIGN" | "N/A"
    """
    publisher = get_value(src, "process.code_signature.subject")
    
    if not publisher:
        return "SUSPICIOUS (Publisher unknown or empty)"
    
    publisher = publisher.lower()
    
    if "microsoft" in publisher:
        return f"BENIGN (Microsoft-signed: {get_value(src, 'process.code_signature.subject')})"
    else:
        return f"MALICIOUS (Non-Microsoft publisher: {get_value(src, 'process.code_signature.subject')})"

def validate_rule9_file_path_legitimacy(src):
    """
    VALIDATION 2 for Rule 9: Check if binary is in legitimate system path
    - Extracts path from both process.executable and process.command_line
    - Mismatch between system32 and syswow64 → SUSPICIOUS (potential evasion)
    - Outside System32/SysWOW64 → MALICIOUS (with both paths shown)
    - Paths differ significantly → MALICIOUS (masquerading attempt)
    - In System32/SysWOW64 → BENIGN
    
    Returns: "MALICIOUS (Exec: C:\\... | Cmd: C:\\...)" | "SUSPICIOUS (Path mismatch)" | "BENIGN (Path: C:\\...)" | "N/A"
    """
    proc_executable = get_value(src, "process.executable")
    proc_cmdline = get_value(src, "process.command_line")
    
    if not proc_executable:
        return "N/A (No executable path)"
    
    proc_executable_lower = proc_executable.lower()
    system_paths = ["system32", "syswow64"]
    
    # Extract path from command line (first token, usually quoted or unquoted path)
    cmdline_path = ""
    if proc_cmdline:
        # Remove quotes and get first token (the executable path)
        first_token = proc_cmdline.split()[0].strip('"\'')
        cmdline_path = first_token
    
    # Check if executable is in legitimate system paths
    in_system = any(sys_path in proc_executable_lower for sys_path in system_paths)
    
    # Check if command line path differs from executable path (potential masquerading)
    paths_differ = False
    if cmdline_path and cmdline_path.lower() != proc_executable_lower:
        paths_differ = True
    
    # Check if mismatch is between system32 and syswow64 (flag as SUSPICIOUS)
    system32_syswow64_mismatch = False
    if paths_differ and cmdline_path:
        cmdline_path_lower = cmdline_path.lower()
        has_system32_exec = "system32" in proc_executable_lower
        has_syswow64_exec = "syswow64" in proc_executable_lower
        has_system32_cmd = "system32" in cmdline_path_lower
        has_syswow64_cmd = "syswow64" in cmdline_path_lower
        
        # If one has system32 and other has syswow64, flag as SUSPICIOUS
        if (has_system32_exec and has_syswow64_cmd) or (has_syswow64_exec and has_system32_cmd):
            system32_syswow64_mismatch = True
    
    if system32_syswow64_mismatch:
        return f"SUSPICIOUS (System32/SysWOW64 mismatch - Exec: {proc_executable} | Cmd: {cmdline_path})"
    
    if paths_differ:
        return f"MALICIOUS (Path mismatch - Exec: {proc_executable} | Cmd: {cmdline_path})"
    
    if in_system:
        return f"BENIGN (System path: {proc_executable})"
    else:
        return f"MALICIOUS (Non-system path: {proc_executable})"

def validate_rule10_parent_process(src):
    """
    VALIDATION 1 for Rule 10: Check parent process for SMB remote execution
    - Suspicious parent (Office/browsers/scripts) → MALICIOUS (infection vector)
    - Legitimate installer parent → SUSPICIOUS (uncommon but possible)
    - No parent info → MALICIOUS
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    parent_name = get_value(src, "process.parent.name")
    
    if not parent_name:
        return "MALICIOUS (No parent process information)"
    
    parent_name_lower = parent_name.lower()
    
    # Check for suspicious parents (infection vectors)
    if any(susp in parent_name_lower for susp in SUSPICIOUS_PARENTS):
        return f"MALICIOUS (Suspicious parent: {parent_name})"
    
    # Check for legitimate installer parents
    if any(legit in parent_name_lower for legit in LEGITIMATE_PARENTS):
        return f"SUSPICIOUS (Legit parent: {parent_name} - Uncommon approach for SMB DLL registration)"
    
    # Unknown parent
    return f"SUSPICIOUS (Unknown parent: {parent_name})"

def validate_rule10_network_context(src):
    """
    VALIDATION 2 for Rule 10: Analyze network context (internal vs external)
    - External network → MALICIOUS (C2 communication)
    - Internal network + normal user → MALICIOUS (lateral movement)
    - Internal network + admin/system → SUSPICIOUS (uncommon for DLL registration)
    
    Returns: "MALICIOUS" | "SUSPICIOUS" | "N/A"
    """
    dest_ip = get_value(src, "destination.ip")
    user_id = get_value(src, "user.id")
    user_name = get_value(src, "user.name")
    
    if not dest_ip:
        return "N/A (No destination IP)"
    
    # Check if destination IP is internal (RFC1918 private ranges)
    is_internal = (
        dest_ip.startswith("10.") or
        dest_ip.startswith("172.16.") or dest_ip.startswith("172.17.") or 
        dest_ip.startswith("172.18.") or dest_ip.startswith("172.19.") or
        dest_ip.startswith("172.20.") or dest_ip.startswith("172.21.") or
        dest_ip.startswith("172.22.") or dest_ip.startswith("172.23.") or
        dest_ip.startswith("172.24.") or dest_ip.startswith("172.25.") or
        dest_ip.startswith("172.26.") or dest_ip.startswith("172.27.") or
        dest_ip.startswith("172.28.") or dest_ip.startswith("172.29.") or
        dest_ip.startswith("172.30.") or dest_ip.startswith("172.31.") or
        dest_ip.startswith("192.168.") or
        dest_ip.startswith("127.") or  # Loopback
        dest_ip == "localhost"
    )
    
    if not is_internal:
        return f"MALICIOUS (External SMB connection: {dest_ip})"
    
    # Internal network - check user privilege
    privilege = check_user_privilege(user_name, user_id)
    
    if privilege in ["admin", "system"]:
        return f"SUSPICIOUS ({privilege} user, internal SMB: {dest_ip} - Uncommon for DLL registration)"
    else:
        return f"MALICIOUS (Normal user lateral movement: {dest_ip})"

# ---------------------------------------------------------------------------
# Per-rule detection queries and aggregation
# ---------------------------------------------------------------------------

def build_rule_query(rule_num, time_range=None):
    """Return an Elasticsearch query body for a single rule (1-11).
    These map directly to the rule descriptions shown in documentation.
    """
    base_must = [
        {"term": {"host.os.type": "windows"}},
        {"term": {"event.type": "start"}},
        {"match": {"process.name": "regsvr32.exe"}},
    ]

    if rule_num == 1:
        # Rule 1: Scriptlet/Remote execution - requires /i or -i flag AND suspicious indicators
        # Returns ALL matches - validation columns will classify by user context and parent process
        # User/parent validation happens in Python (validate_rule1_user_context, validate_rule1_parent_process)
        # NOT in Elasticsearch query - this just returns all matching patterns
        rule = {
            "bool": {
                "must": base_must + [
                    {
                        "bool": {
                            "should": [
                                {"wildcard": {"process.command_line": "*/i*"}},
                                {"wildcard": {"process.command_line": "*-i*"}}
                            ],
                            "minimum_should_match": 1
                        }
                    },
                    {
                        "bool": {
                            "should": [
                                {"wildcard": {"process.command_line": "*scrobj.dll*"}},
                                {"wildcard": {"process.command_line": "*http*"}},
                                {"wildcard": {"process.command_line": "*\\\\\\\\*"}},  # UNC path (double backslash)
                                {"wildcard": {"process.command_line": "*ftp*"}},
                                {"wildcard": {"process.command_line": "*.sct*"}}
                            ],
                            "minimum_should_match": 1
                        }
                    }
                ]
              
            }
        }
    elif rule_num == 2:
        # Rule 2: Non-system path DLL - regsvr32 from temp/user directories
        paths = ["*C:\\Users\\*", "*C:\\ProgramData\\*", "*C:\\Windows\\Temp\\*", "*\\AppData\\Local\\Temp\\*", "*C:\\Users\\Public\\*", "*\\Downloads\\*", "*\\Desktop\\*"]
        should_paths = [{"wildcard": {"process.command_line": p}} for p in paths]
        rule = {"bool": {"must": base_must, "should": should_paths, "minimum_should_match": 1}}
    elif rule_num == 3:
        # Rule 3: Non-standard extension - EXE/PS1/BAT/VBS/etc instead of DLL
        exts = ["*.exe","*.ps1","*.bat","*.cmd","*.vbs","*.js","*.txt","*.pdf","*.doc","*.docx","*.xls","*.xlsx","*.zip"]
        must_ext = [{"wildcard": {"process.command_line": e}} for e in exts]
        must_not = [{"wildcard": {"process.command_line": "*.dll*"}}, {"wildcard": {"process.command_line": "*.ocx*"}}, {"wildcard": {"process.command_line": "*.ax*"}}]
        rule = {"bool": {"must": base_must + must_ext, "must_not": must_not}}
    elif rule_num == 4:
        # Rule 4: Double-extension masquerade - PDF,DLL, DOC.DLL, etc.
        double_exts = ["*.pdf.*","*.doc.*","*.docx.*","*.xls.*","*.xlsx.*","*.ppt.*","*.jpg.*","*.png.*"]
        second = ["*.dll*","*.sct*","*.ocx*","*.cpl*"]
        should_clauses = []
        for d in double_exts:
            for s in second:
                should_clauses.append({"bool": {"must": [{"wildcard": {"process.command_line": d}}, {"wildcard": {"process.command_line": s}}]}})
        rule = {"bool": {"must": base_must, "should": should_clauses, "minimum_should_match": 1}}
    elif rule_num == 5:
        # Rule 5: Network activity - regsvr32 making network connections
        rule = {"bool": {"must": base_must + [{"term": {"event.category": "network"}}]}}
    elif rule_num == 6:
        # Rule 6: Suspicious parent - Office/Browser/PowerShell/CMD/WScript/WMI
        # regsvr32.exe spawned by Office apps, browsers, shells, or scripting engines
        # Also detects admin-level evasion: Task Scheduler, services, WMI execution
        rule = {"bool": {"must": base_must + [{"terms": {"process.parent.name": SUSPICIOUS_PARENTS}}]}}
    elif rule_num == 7:
        # Rule 7: DLL unsigned - no digital signature (unsigned DLL loads)
        rule = {"bool": {"must": base_must + [{"term": {"event.category": "library"}}, {"term": {"dll.code_signature.exists": False}}]}}
    elif rule_num == 8:
        # Rule 8: Child processes spawned by regsvr32
        rule = {"bool": {"must": [{"term": {"host.os.type": "windows"}}, {"term": {"event.type": "start"}}, {"match": {"process.parent.name": "regsvr32.exe"}}]}}
    elif rule_num == 9:
        # Rule 9: regsvr32 process unsigned or untrusted signature
        rule = {"bool": {"must": base_must, "should": [{"term": {"process.code_signature.exists": False}}, {"term": {"process.code_signature.trusted": False}}], "minimum_should_match": 1}}
    elif rule_num == 10:
        # Rule 10: SMB Share Remote Execution - UNC path + SMB network (ports 445/139)
        rule = {
            "bool": {
                "must": base_must + [
                    {"wildcard": {"process.command_line": "*\\\\*"}},
                    {"term": {"event.category": "network"}}
                ],
                # "should": [
                #     {"term": {"destination.port": 445}},
                #     {"term": {"destination.port": 139}}
                # ],
                "minimum_should_match": 1
            }
        }
    elif rule_num == 11:
        # Rule 11: Renamed regsvr32 - original filename mismatch
        rule = {"bool": {"must": base_must + [{"match": {"process.pe.original_file_name": "REGSVR32.EXE"}}], "must_not": [{"match": {"process.name": "regsvr32.exe"}}]}}
    # elif rule_num == 12:
    #     # Rule 12: DLL from non-legitimate path (anything OTHER than system paths/program files)
    #     # Detects DLLs from user-writable paths, temp directories, removable media, network shares, etc.
    #     # Uses case-insensitive wildcards to match Windows paths (C:\Program Files vs C:\program files)
    #     legit_paths = [
    #         "*\\\\windows\\\\system32\\\\*",
    #         "*\\\\windows\\\\syswow64\\\\*",
    #         "*\\\\program files\\\\*",
    #         "*\\\\program files (x86)\\\\*",
    #         "*\\\\Windows\\\\System32\\\\*",  # Case variations
    #         "*\\\\Windows\\\\SysWOW64\\\\*",
    #         "*\\\\Program Files\\\\*",
    #         "*\\\\Program Files (x86)\\\\*"
    #     ]
    #     must_not = [{"wildcard": {"dll.path": {"value": path, "case_insensitive": True}}} for path in legit_paths[:4]]
    #     rule = {"bool": {"must": base_must + [{"term": {"event.category": "library"}}], "must_not": must_not}}
    else:
        raise ValueError("Unknown rule number")

   
    if time_range:
        # Add time filter into the rule's bool->filter list instead of wrapping the entire rule
        rule_bool = rule.get("bool", {})
        rule_bool.setdefault("filter", [])
        rule_bool["filter"].append({"range": {"@timestamp": {"gte": time_range}}})
        rule["bool"] = rule_bool
        return rule

    # Return the rule (a dict with 'bool')
    return rule

def fetch_all_scroll(es_client, index, query_body, size=5000, scroll_ttl="2m", source_fields=None):
    """Fetch all hits using scroll API.

    This function is defensive: callers may pass either the inner rule (e.g. {"bool": {...}})
    or a full search body (e.g. {"query": {...}}). We normalize to a full body and log
    helpful debugging info on error to diagnose BadRequest errors from Elasticsearch.
    """
    all_hits = []
    # Normalize body: if caller already passed a full body (has 'query' key), use it directly
    if isinstance(query_body, dict) and "query" in query_body:
        body = query_body
    else:
        body = {"query": query_body}
    
    # Add size and _source to body to avoid deprecation warning
    body["size"] = size
    if source_fields:
        body["_source"] = source_fields

    try:
        resp = es_client.search(index=index, body=body, scroll=scroll_ttl)
    except Exception as e:
        try:
            import json
            printed = json.dumps(body, indent=2)[:2000]
        except Exception:
            printed = str(body)
        print(f"[!] Initial search failed: {e}\nRequest body (truncated): {printed}")
        return all_hits

    scroll_id = resp.get("_scroll_id")
    hits = resp.get("hits", {}).get("hits", [])
    all_hits.extend(hits)

    while hits:
        try:
            resp = es_client.scroll(scroll_id=scroll_id, scroll=scroll_ttl)
        except Exception as e:
            print(f"[!] Error during scroll: {e}")
            break
        scroll_id = resp.get("_scroll_id")
        hits = resp.get("hits", {}).get("hits", [])
        if not hits:
            break
        all_hits.extend(hits)

    try:
        if scroll_id:
            es_client.clear_scroll(scroll_id=scroll_id)
    except Exception:
        pass

    return all_hits

def build_general_regsvr_query(time_range):
    """
    Build a general query for regsvr32 detection (Windows, event.type=start).
    No strict filtering - just process name matching.
    """
    query = {
        "bool": {
            "must": [
                {"term": {"host.os.type": "windows"}},
                {"term": {"event.type": "start"}},
                {"match": {"process.name": "regsvr32.exe"}},
            ],
            "filter": []
        }
    }
    
    if time_range:
        query["bool"]["filter"].append({"range": {"@timestamp": {"gte": time_range}}})
    
    return query

def query_all_rules(es_client, time_range=None):
    """Run all per-rule queries (1-11) and aggregate unique hits with a 'matched_rules' field in _source.
    Returns a list of unique hits (each hit['_source']['matched_rules'] is a string with rule IDs and labels).
    Rule 12 is commented out and excluded from export.
    """
    aggregated = {}
    for i in range(1, 12):
        print(f"[+] Running Rule {i} query...")
        q = build_rule_query(i, time_range=time_range)
        hits = fetch_all_scroll(es_client, ES_INDEX, q, size=5000, scroll_ttl="2m", source_fields=SOURCE_FIELDS)
        print(f"    - Rule {i}: {len(hits)} hits")
        for h in hits:
            hid = h.get("_id")
            if hid is None:
                # Fallback: use a deterministic key composed of timestamp+host+process.pid
                src = h.get("_source", {})
                hid = f"{src.get('@timestamp','')}-{src.get('host', {}).get('name','')}-{src.get('process',{}).get('pid','')}-{len(aggregated)+1}"
            if hid not in aggregated:
                # copy hit to avoid mutating shared object
                new_hit = dict(h)
                new_hit['_source'] = dict(h.get('_source', {}))
                new_hit['_source']['matched_rules'] = []
                aggregated[hid] = new_hit
            # Append rule id + short description
            aggregated[hid]['_source']['matched_rules'].append(f"{i}: {RULE_DESCRIPTIONS.get(i,'')}")

    # Convert matched_rules list to comma-separated string
    final_hits = []
    for hid, hit in aggregated.items():
        hit['_source']['matched_rules'] = ', '.join(hit['_source'].get('matched_rules', []))
        final_hits.append(hit)

    print(f"[+] Aggregated {len(final_hits)} unique hits across all rules")
    return final_hits

def query_rules_separately(es_client, time_range=None):
    """Run per-rule queries (1-11) and return a dict of rule_num -> hits list.
    Rule 12 is commented out and excluded from export.
    """
    rule_hits = {}
    for i in range(1, 12):
        print(f"[+] Running Rule {i} query (separate sheet)...")
        q = build_rule_query(i, time_range=time_range)
        hits = fetch_all_scroll(es_client, ES_INDEX, q, size=5000, scroll_ttl="2m", source_fields=SOURCE_FIELDS)
        print(f"    - Rule {i}: {len(hits)} hits")
        rule_hits[i] = hits
    return rule_hits

def add_sheet_with_data(wb, hits, sheet_title, header_color="C00000", field_list=None, description=None, is_rule1=False):
    """Add a worksheet with data to an existing workbook.
    
    Args:
        wb: Workbook object
        hits: List of Elasticsearch hits
        sheet_title: Title of the sheet
        header_color: Header background color (hex)
        field_list: Custom field list (defaults to FIELDS if None)
        description: Optional description string to place above the header row
        is_rule1: True if this is Rule 1 sheet (adds validation columns)
    """
    ws = wb.create_sheet(title=sheet_title[:31])
    
    # Use custom field list or default FIELDS
    if field_list is None:
        field_list = FIELDS

    # Optional description row
    start_row = 1
    if description:
        ws.append([description])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(field_list))
        desc_cell = ws.cell(row=1, column=1)
        desc_cell.font = Font(bold=True, color="000000")
        desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        start_row = 2
    
    # Header with styling
    ws.append(field_list)
    fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    header_row = start_row
    for col_idx in range(1, len(field_list) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # Auto-size columns - wider for validation columns
        if is_rule1 and col_idx <= 2:
            ws.column_dimensions[get_column_letter(col_idx)].width = 60  # Wider for validation columns
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Add data rows with pattern matching
    for hit in hits:
        src = hit.get('_source', {})
        row = []
        for field in field_list:
            if field == "validation_1_user_context":
                # Rule 1 Validation 1: User context validation
                row.append(validate_rule1_user_context(src))
            elif field == "validation_2_parent_process" and is_rule1:
                # Rule 1 Validation 2: Parent process validation
                row.append(validate_rule1_parent_process(src))
            elif field == "validation_1_parent_process" and not is_rule1:
                # Rule 2 Validation 1: Parent process and child process tree validation
                row.append(validate_rule2_parent_process(src))
            elif field == "validation_2_dll_signature" and not is_rule1:
                # Rule 2 Validation 2: DLL signature validation
                row.append(validate_rule2_dll_signature(src))
            elif field == "Val1 DLL Load Presence":
                # Rule 3 Validation 1: Non-DLL extension + no DLL load
                row.append(validate_rule3_no_dll_load(src))
            elif field == "Val1 File Signature Check":
                # Rule 4 Validation 1: File signature check
                row.append(validate_rule4_signature_check(src))
            elif field == "Val2 DLL Load Detection":
                # Rule 4 Validation 2: DLL load detection
                row.append(validate_rule4_dll_load_detection(src))
            elif field == "Val1 Network Type":
                # Rule 5 Validation 1: Network type (internal vs external)
                row.append(validate_rule5_network_type(src))
            elif field == "Val2 Network Timing":
                # Rule 5 Validation 2: Network activity timing
                row.append(validate_rule5_network_timing(src))
            elif field == "Val1 User Context":
                # Rule 6 Validation 1: User privilege assessment
                row.append(validate_rule6_user_privilege(src))
            elif field == "Val2 Network Connection":
                row.append(validate_rule6_network_connection(src))
            elif field == "Val3 Parent Signature":
                # Rule 6 Validation 3: Parent process signature check
                row.append(validate_rule6_parent_signature(src))
            elif field == "Val1 DLL File Location":
                # Rule 7 Validation 1: DLL file location check
                row.append(validate_rule7_dll_file_location(src))
            elif field == "Val2 Network Activity":
                # Rule 7 Validation 2: Network activity presence after execution
                row.append(validate_rule7_network_activity_presence(src))
            elif field == "Val1 Child Process Risk":
                # Rule 8 Validation 1: Child process type and risk assessment
                row.append(validate_rule8_child_process(src))
            elif field == "Val2 DLL Path in Command":
                # Rule 8 Validation 2: DLL path in command line legitimacy
                row.append(validate_rule8_dll_path_in_command(src))
            elif field == "Val1 Signature Publisher":
                # Rule 9 Validation 1: Code signature publisher identification
                row.append(validate_rule9_signature_publisher(src))
            elif field == "Val2 File Path":
                # Rule 9 Validation 2: File path legitimacy
                row.append(validate_rule9_file_path_legitimacy(src))
            elif field == "Val1 Parent Process":
                # Rule 10 Validation 1: Parent process analysis
                row.append(validate_rule10_parent_process(src))
            elif field == "Val2 Network Context":
                # Rule 10 Validation 2: Network context (internal/external + user privilege)
                row.append(validate_rule10_network_context(src))
            else:
                row.append(serialize_cell(get_value(src, field)))
        ws.append(row)
    
    return len(hits)

def export_hexa_sheets(general_hits, out_path, all_rules_hits=None, per_rule_hits=None, query_only=False, validation_only=False):
    """Export general hits and per-rule data to separate sheets.
    Includes General sheet, Per_Rule_Detections sheet, and separate sheets per rule.
    
    Args:
        general_hits: General regsvr32 events
        out_path: Output Excel file path
        all_rules_hits: Aggregated rule hits for per-rule classification
        per_rule_hits: Dict of rule_num -> hits for per-rule sheets
        query_only: If True, only export 1st Query sheets (skip validation sheets)
        validation_only: If True, only export Validation sheets (skip 1st Query sheets)
    """
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    
    # Add General sheet (blue header) - all events (always included unless validation_only)
    if not validation_only:
        general_count = add_sheet_with_data(wb, general_hits, "General_All_Regsvr32", header_color="4472C4")
    else:
        general_count = 0

    # Add Per-Rule classification sheet (light blue header) if provided (skip if query_only or validation_only)
    per_rule_count = 0
    if all_rules_hits is not None and not query_only and not validation_only:
        per_rule_count = add_sheet_with_data(wb, all_rules_hits, "Per_Rule_Detections", header_color="00B0F0", field_list=RULES_FIELDS)
        print(f"    - Per-rule classification sheet: {per_rule_count} rows")

    # Add separate sheets per rule (one sheet per rule query) if provided
    per_rule_sheet_counts = {}
    if per_rule_hits:
        for rule_num in range(1, 12):
            hits = per_rule_hits.get(rule_num, [])
            sheet_title = f"Rule_{rule_num:02d}"  # Excel sheet limit 31 chars
            description = RULE_DESCRIPTIONS.get(rule_num, "")
            
            # Use special field list and flag for Rule 1 (includes validation columns)
            if rule_num == 1:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 1 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 1 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE1_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 1 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE1_FIELDS,
                        is_rule1=True
                    )
                    print(f"    Rule 1 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 2 (includes validation columns)
            elif rule_num == 2:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 2 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 2 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE2_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 2 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE2_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 2 Validation sheet: {validation_count} rows")
            # Use special field list for Rule 3 (includes validation columns)
            elif rule_num == 3:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 3 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 3 Query sheet: {query_count} rows")

                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE3_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 3 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE3_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 3 Validation sheet: {validation_count} rows")
            # Use special field list for Rule 4 (includes validation columns for double-extension detection)
            elif rule_num == 4:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 4 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 4 Query sheet: {query_count} rows")

                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE4_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 4 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE4_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 4 Validation sheet: {validation_count} rows")
            elif rule_num == 5:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 5 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 5 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE5_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 5 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE5_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 5 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 6 (includes validation columns for user privilege analysis)
            elif rule_num == 6:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 6 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 6 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE6_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 6 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE6_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 6 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 7 (includes validation columns for DLL location and network activity analysis)
            elif rule_num == 7:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 7 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 7 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE7_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 7 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE7_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 7 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 8 (includes validation columns for child process analysis)
            elif rule_num == 8:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 8 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 8 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE8_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 8 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE8_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 8 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 9 (includes validation columns for code signature analysis)
            elif rule_num == 9:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 9 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 9 Query sheet: {query_count} rows")
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE9_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 9 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE9_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 9 Validation sheet: {validation_count} rows")
            # Use special field list and flag for Rule 10 (includes validation columns for SMB remote execution)
            elif rule_num == 10:
                if not validation_only:
                    # Sheet 1: Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 10 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule 10 Query sheet: {query_count} rows")
                
                if not query_only:
                    # Sheet 2: Validation results with full telemetry (RULE10_FIELDS)
                    validation_count = add_sheet_with_data(
                        wb,
                        hits,
                        "Rule 10 Validation",
                        header_color="70AD47",
                        description=f"Rule {rule_num} Validation: {description}" if description else f"Rule {rule_num} Validation",
                        field_list=RULE10_FIELDS,
                        is_rule1=False
                    )
                    print(f"    Rule 10 Validation sheet: {validation_count} rows")
            else:
                # Rules 3, 4, 11, 12: No custom validation columns
                if not validation_only:
                    # Query results only (SOURCE_FIELDS)
                    query_count = add_sheet_with_data(
                        wb,
                        hits,
                        f"Rule {rule_num} 1st Query",
                        header_color="5B9BD5",
                        description=f"Rule {rule_num}: {description}" if description else f"Rule {rule_num}",
                        field_list=SOURCE_FIELDS
                    )
                    print(f"    Rule {rule_num} Query sheet: {query_count} rows")
                    per_rule_sheet_counts[rule_num] = query_count

    
    # Remove the default empty sheet
    wb.remove(default_sheet)
    
    # Save workbook
    wb.save(out_path)
    
    return general_count, per_rule_count

def workflow_import_from_elastic_cli(time_range, dry_run=False, validate=False, query_only=False, validation_only=False):
    """CLI version: Import regsvr32 data with 6 sheets (includes high-risk, medium-risk and low-risk detection).

    If dry_run is True: print the request bodies for general/specific/child/high-risk queries and for per-rule queries
    (1-11) and exit without contacting Elasticsearch.

    If validate is True: connect to Elasticsearch and call the indices.validate_query API for each query body
    and print validation response (helpful for diagnosing BadRequest errors from invalid query bodies).
    """
    print("\n[+] Running Workflow 1: Import regsvr32 data (6 sheets with risk-based filtering)")

    if dry_run:
        import json
        print("[i] Dry-run mode: printing query bodies (no requests will be sent to Elasticsearch)")

        # General query
        print("\n--- General Query ---")
        gen_q = {"query": build_general_regsvr_query(time_range)}
        print(json.dumps(gen_q, indent=2)[:4000])



        # Per-rule queries 1-11
        print("\n--- Per-Rule Queries (1-11) ---")
        for i in range(1, 12):
            try:
                q = build_rule_query(i, time_range=time_range)
            except Exception as e:
                print(f"[!] Failed to build Rule {i} query: {e}")
                continue
            body = {"query": q} if not (isinstance(q, dict) and "query" in q) else q
            print(f"\nRule {i} request body:")
            print(json.dumps(body, indent=2)[:4000])
        print("\n[✓] Dry-run complete. No requests were sent.")
        return

    # If validation-only mode requested, connect to ES and validate each body
    if validate:
        print("[+] Validation mode: connecting to Elasticsearch to validate queries...")
        es = Elasticsearch([ES_URL], api_key=ES_API_KEY, verify_certs=VERIFY_CERTS)
        try:
            import json
            # General
            gen_body = {"query": build_general_regsvr_query(time_range)}
            print("\n--- Validate General Query ---")
            resp = es.indices.validate_query(index=ES_INDEX, body=gen_body, explain=True)
            print(json.dumps(resp, indent=2)[:4000])

        

            # Per-rule validations
            print("\n--- Validate Per-Rule Queries (1-11) ---")
            for i in range(1, 12):
                try:
                    q = build_rule_query(i, time_range=time_range)
                except Exception as e:
                    print(f"[!] Failed to build Rule {i} query: {e}")
                    continue
                body = {"query": q} if not (isinstance(q, dict) and "query" in q) else q
                print(f"\nValidate Rule {i} request body:")
                print(json.dumps(body, indent=2)[:4000])
                try:
                    resp = es.indices.validate_query(index=ES_INDEX, body=body, explain=True)
                    print(json.dumps(resp, indent=2)[:4000])
                except Exception as e:
                    print(f"[!] Validation for Rule {i} failed: {e}")
        except Exception as e:
            print(f"[!] Validation run failed: {e}")
        return

    # Connect to Elasticsearch
    print("[+] Connecting to Elasticsearch...")
    es = Elasticsearch([ES_URL], api_key=ES_API_KEY, verify_certs=VERIFY_CERTS)

    # Query 1: General (all regsvr32)
    print(f"[+] Querying for regsvr32.exe (general - all events)...")
    general_query = build_general_regsvr_query(time_range)
    general_hits = fetch_all_scroll(es, ES_INDEX, general_query, size=5000, scroll_ttl="2m", source_fields=SOURCE_FIELDS)
    
    if not general_hits:
        print(f"[-] No regsvr32.exe events found in the specified time range")
        return
    
    print(f"[+] Retrieved {len(general_hits)} general events")
    
    # Query per-rule classification (run each rule 1-11 and aggregate)
    print(f"[+] Running per-rule queries (1-11) and aggregating results for classification...")
    all_rules_hits = query_all_rules(es, time_range=time_range)

    # Per-rule hits for separate sheets (one sheet per rule)
    print(f"[+] Running per-rule queries (1-11) for separate sheets...")
    per_rule_hits = query_rules_separately(es, time_range=time_range)

    # Generate output filename
    ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    time_label = "custom" if time_range else "all_time"
    out_fn = f"regsvr32_perrulerule_{time_label}_{ts}.xlsx"
    out_path = os.path.join(".", out_fn)
    
    # Export to Excel with per-rule sheets
    export_hexa_sheets(
        general_hits,
        out_path,
        all_rules_hits=all_rules_hits,
        per_rule_hits=per_rule_hits,
        query_only=query_only,
        validation_only=validation_only,
    )
    print(f"\n[✓] Export complete! File saved: {out_path}")
    if validation_only:
        print(f"[i] Sheets: Validation columns only (no 1st Query sheets)")
    elif query_only:
        print(f"[i] Sheets: 1st Query results only (no Validation sheets)")
    else:
        print(f"[i] Sheet 1 (Blue): All regsvr32 events")
        print(f"[i] Sheet 2 (Light Blue): Per-rule classification")
        print(f"[i] Additional sheets: Rule_01 ... Rule_11 (Query + Validation for each rule)")

def main():
    parser = argparse.ArgumentParser(
        description="Regsvr32 Threat Hunting - 6 Sheet Exporter with High-Risk Detection",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Import all regsvr32 data (6 sheets) from last 24 hours
    python regsvr32_hunter.py import 24h

  # Import from last 7 days
    python regsvr32_hunter.py import 7d

  # Import from last 1 month
    python regsvr32_hunter.py import 1mo

  # Import from last 6 months
    python regsvr32_hunter.py import 6mo

  # Import from last 1 year
    python regsvr32_hunter.py import 1y

  # Import all time data
    python regsvr32_hunter.py import all

Time format examples:
  10s, 30s, 5m, 30m, 1h, 6h, 12h, 24h, 2d, 7d, 30d
  1mo, 2mo, 3mo, 6mo (months)
  1y, 2y (years)
  all (all time)

        """
    )
    
    parser.add_argument(
        "workflow",
        choices=["import"],
        help="Workflow to run: import (6 sheets with isolated high-risk sheet + AWL bypass pattern)"
    )
    
    parser.add_argument(
        "time",
        help="Time range: 10s, 30s, 5m, 30m, 1h, 6h, 12h, 24h, 2d, 7d, 30d, 1mo, 2mo, 3mo, 6mo, 1y, 2y, all"
    )
    parser.add_argument("--dry-run", action="store_true", help="Print query bodies (general, specific and per-rule) and exit without sending requests to Elasticsearch")
    parser.add_argument("--validate", action="store_true", help="Validate queries using Elasticsearch validate API (requires ES access)")
    parser.add_argument("--query-only", action="store_true", help="Only export 1st Query sheets (skip Validation sheets)")
    parser.add_argument("--validation-only", action="store_true", help="Only export Validation sheets (skip 1st Query sheets)")
    
    args = parser.parse_args()
    
    # Parse time range
    time_str = args.time.lower().strip()
    if time_str == "all":
        time_range = None
        time_label = "all_time"
    else:
        # Convert shorthand to Elasticsearch format
        import re
        
        # Match seconds, minutes, hours, days
        match = re.match(r"(\d+)([smhd])$", time_str)
        if match:
            num, unit = match.groups()
            num = int(num)
            
            # Convert days to hours for Elasticsearch
            if unit == "d":
                num = num * 24
                unit = "h"
            
            time_range = f"now-{num}{unit}"
            time_label = time_str
        else:
            # Match months (mo) or years (y)
            match = re.match(r"(\d+)(mo|y)$", time_str)
            if match:
                num, unit = match.groups()
                num = int(num)
                
                # Convert to hours for Elasticsearch
                if unit == "mo":
                    num = num * 30 * 24  # 30 days per month
                    unit = "h"
                    time_label = f"{num//24//30}mo"
                elif unit == "y":
                    num = num * 365 * 24  # 365 days per year
                    unit = "h"
                    time_label = f"{num//24//365}y"
                
                time_range = f"now-{num}{unit}"
            else:
                print(f"[!] Invalid time format: {time_str}")
                print("[!] Use format like: 10s, 5m, 1h, 7d, 1mo, 6mo, 1y, or 'all'")
                return
    
    print("\n" + "="*70)
    print("REGSVR32 THREAT HUNTING - 6 SHEET EXPORTER (HIGH-RISK DETECTION)")
    print("="*70)
    print(f"\nWorkflow: {args.workflow}")
    print(f"Time range: {time_label}")
    if args.query_only:
        print(f"Output mode: 1st Query sheets ONLY")
    elif args.validation_only:
        print(f"Output mode: Validation sheets ONLY")
    print("="*70)
    
    # Validate conflicting options
    if args.query_only and args.validation_only:
        print("[!] Error: Cannot use --query-only and --validation-only together")
        return
    
    # Run workflow
    if args.workflow == "import":
        workflow_import_from_elastic_cli(
            time_range, 
            dry_run=args.dry_run,
            query_only=args.query_only,
            validation_only=args.validation_only
        )
    
    print("\n[✓] Done!")

if __name__ == '__main__':
    main()
